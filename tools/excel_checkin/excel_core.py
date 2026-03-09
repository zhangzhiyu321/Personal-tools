"""
Excel 打卡标红工具的核心逻辑（与具体 Web/CLI 入口解耦）。

后端 Web 入口和命令行入口都应调用这里的函数，避免重复实现。
"""

import os
import re
import traceback
from datetime import datetime, time, timedelta
from typing import List, Dict, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from .log import get_logger


logger = get_logger()
RED_FONT = Font(color="FFFF0000")

# 清洗打卡时间时保留的关键词（如「外勤」），会按原顺序追加到单元格末尾并换行显示，可扩展
PRESERVE_KEYWORDS = ("外勤",)


def is_weekend(date_str: str) -> bool:
    """判断是否是周末（周六、周日）"""
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        weekday = date_obj.weekday()
        return weekday >= 5
    except Exception:
        return False


def extract_last_time(time_str) -> Optional[time]:
    """从打卡时间字符串中提取最后一次打卡时间"""
    if pd.isna(time_str) or time_str == "":
        return None

    try:
        import re

        time_str = str(time_str).strip()
        if not time_str:
            return None

        if "休息" in time_str:
            return None

        # 注意：这里是普通正则，匹配 1-2 位数字:1-2 位数字
        # 之前写成 "(\\d{1,2}):(\\d{1,2})" 会导致正则变成匹配反斜杠+数字，无法识别时间
        time_pattern = r"(\d{1,2}):(\d{1,2})"
        matches = re.findall(time_pattern, time_str)

        if not matches:
            return None

        hour_str, minute_str = matches[-1]
        hour = int(hour_str)
        minute = int(minute_str)

        if hour < 0 or hour > 23 or minute < 0 or minute > 59:
            return None

        return time(hour, minute)
    except Exception:
        return None


def extract_time_points(time_value) -> List[time]:
    """从单元格内容中提取所有时间点（HH:MM），返回去重后的升序列表。"""
    if pd.isna(time_value) or time_value == "":
        return []

    try:
        import re

        s = str(time_value).strip()
        if not s or "休息" in s:
            return []

        matches = re.findall(r"(\d{1,2}):(\d{1,2})", s)
        points: List[time] = []
        for h_str, m_str in matches:
            try:
                h = int(h_str)
                m = int(m_str)
                if 0 <= h <= 23 and 0 <= m <= 59:
                    points.append(time(h, m))
            except Exception:
                continue

        if not points:
            return []

        # 去重 + 升序（同一天只关心时间点集合）
        uniq = sorted(set(points))
        return uniq
    except Exception:
        return []


def _format_hhmm(t: time) -> str:
    return f"{t.hour:02d}:{t.minute:02d}"


def _append_preserved_keywords(original_text: str, result: str) -> str:
    """若原单元格包含 PRESERVE_KEYWORDS 中的词，按出现顺序追加到 result 后并换行，保证通用可扩展。"""
    if not result or not original_text:
        return result
    s = str(original_text).strip()
    added = []
    for kw in PRESERVE_KEYWORDS:
        if kw and kw in s and kw not in added:
            added.append(kw)
    if not added:
        return result
    return result + "\n" + "\n".join(added)


def clean_daily_punch_cell(time_value, cutoff_time: time) -> Optional[str]:
    """按规则清洗每日打卡单元格内容。

    - 时间点数量 < 2：不处理，返回原字符串（保留换行、外勤等）
    - 第 1 个时间点（最早）> cutoff_time：只保留最后一次打卡时间（最晚）
    - 否则：删除第 1 个与第 2 个时间点之间的所有时间（保留最早与最晚两次）
    - 多个时间之间用换行分隔；若原内容含 PRESERVE_KEYWORDS（如「外勤」），会追加到末尾并换行
    """
    if pd.isna(time_value) or time_value == "":
        return None

    s = str(time_value).strip()
    if not s or "休息" in s:
        return s

    points = extract_time_points(time_value)
    if len(points) < 2:
        return s

    first_time = points[0]
    last_time = points[-1]

    if first_time > cutoff_time:
        result = _format_hhmm(last_time)
        return _append_preserved_keywords(s, result)

    if first_time == last_time:
        result = _format_hhmm(first_time)
        return _append_preserved_keywords(s, result)

    # 两个时间用换行分隔，便于在 Excel 中一行一个
    result = f"{_format_hhmm(first_time)}\n{_format_hhmm(last_time)}"
    return _append_preserved_keywords(s, result)


def parse_date_from_header(header_value, base_date_str, column_index, first_date_col_index):
    if pd.isna(header_value):
        return None

    header_str = str(header_value).strip()

    try:
        if "-" in header_str and len(header_str) >= 8:
            return header_str[:10]
    except Exception:
        pass

    try:
        if base_date_str:
            base_date = datetime.strptime(base_date_str[:10], "%Y-%m-%d")
        else:
            return None

        days_offset = column_index - first_date_col_index
        result_date = base_date + timedelta(days=days_offset)

        if header_str.isdigit():
            day = int(header_str)
            if result_date.day == day:
                return result_date.strftime("%Y-%m-%d")
            try:
                result_date = base_date.replace(day=day)
                return result_date.strftime("%Y-%m-%d")
            except Exception:
                pass

        weekdays_map = {"一": 0, "二": 1, "三": 2, "四": 3, "五": 4, "六": 5, "日": 6, "天": 6}
        if header_str in weekdays_map:
            target_weekday = weekdays_map[header_str]
            if result_date.weekday() == target_weekday:
                return result_date.strftime("%Y-%m-%d")
            base_weekday = base_date.weekday()
            days_diff = (target_weekday - base_weekday) % 7
            if days_diff == 0 and header_str not in ["六", "日", "天"]:
                days_diff = 7
            result_date = base_date + timedelta(days=days_diff)
            return result_date.strftime("%Y-%m-%d")

        return result_date.strftime("%Y-%m-%d")
    except Exception:
        return None


def process_excel_for_web(
    file_path: str,
    late_time: time,
    early_time: time,
    custom_holidays: Optional[List[Dict]] = None,
    clean_cutoff_time: time = time(17, 44),
) -> Tuple[Optional[str], List[Dict]]:
    """Web 场景下使用的处理函数，返回输出文件路径和凌晨打卡明细。"""
    from datetime import datetime as _dt

    if custom_holidays is None:
        custom_holidays = []

    def is_holiday(date_str: str, custom_holidays_list: List[Dict]) -> bool:
        """判断日期是否是假期。

        规则：
        - 如果前端没有传自定义假期（custom_holidays_list 为空），则周六周日视为假期；
        - 一旦有自定义假期配置，就只按自定义列表判断，周末不再自动当成假期。
        """
        # 有自定义配置：只按自定义来
        if custom_holidays_list:
            for holiday_item in custom_holidays_list:
                try:
                    if holiday_item.get("type") == "single":
                        if date_str == holiday_item.get("date"):
                            return True
                    elif holiday_item.get("type") == "range":
                        start_date = datetime.strptime(holiday_item.get("start"), "%Y-%m-%d")
                        end_date = datetime.strptime(holiday_item.get("end"), "%Y-%m-%d")
                        current_date = datetime.strptime(date_str, "%Y-%m-%d")
                        if start_date <= current_date <= end_date:
                            return True
                except Exception:
                    continue
            return False

        # 无自定义配置：默认周末是假期
        return is_weekend(date_str)

    try:
        logger.info("excel_core_start | path=%s", file_path)
        wb = load_workbook(file_path)
        ws = wb.active
        logger.debug("workbook_loaded | sheet=%s max_row=%s max_column=%s", ws.title, ws.max_row, ws.max_column)

        base_date_str = None
        name_row_idx = None   # 有「姓名」的行（1-based）
        date_header_row_idx = None  # 有星期/日期列名的行（1-based）

        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_values = [cell.value for cell in ws[row_idx]]
            if base_date_str is None:
                for cell_value in row_values:
                    if cell_value and isinstance(cell_value, str) and "统计日期" in cell_value:
                        date_match = re.search(r"(\d{4}-\d{2}-\d{2})", cell_value)
                        if date_match:
                            base_date_str = date_match.group(1)
                            logger.debug("base_date_found | row=%s value=%s", row_idx, base_date_str)
                            break
            row_str = " ".join([str(v) if v else "" for v in row_values])
            has_name = "姓名" in row_str
            has_weekdays = any(day in row_str for day in ["一", "二", "三", "四", "五", "六", "日"])
            # 支持仅数字的日期列名行，如 "2 3 4 5 6 六 日" 中 2,3,4,5,6 与 六、日 一起出现
            if not has_weekdays and row_str.strip():
                tokens = [str(v).strip() for v in row_values if v is not None and str(v).strip()]
                digit_tokens = [t for t in tokens if t.isdigit() and len(t) <= 2]
                if len(digit_tokens) >= 3:
                    has_weekdays = True
            if has_name and name_row_idx is None:
                name_row_idx = row_idx
            # 排除标题行（含「统计日期」）和报表生成行，避免「日」字被误判为星期列
            is_title_or_report_row = "统计日期" in row_str or "报表生成" in row_str
            if has_weekdays and date_header_row_idx is None and not is_title_or_report_row:
                date_header_row_idx = row_idx
            logger.debug(
                "header_scan_row | row_idx=%s has_name=%s has_weekdays=%s row_preview=%s",
                row_idx,
                has_name,
                has_weekdays,
                (row_str[:120] + "..." if len(row_str) > 120 else row_str),
            )

        # 确定表头行：支持单行表头 或 两行表头（姓名等一行 + 日期列一行）
        header_row_idx = None
        use_two_row_header = False
        if date_header_row_idx is None:
            logger.error(
                "excel_core_fail | reason=date_header_row_not_found scanned_rows=1..%s base_date_str=%s",
                min(10, ws.max_row),
                base_date_str,
            )
            return None, []
        header_row_idx = date_header_row_idx
        if name_row_idx is not None and date_header_row_idx == name_row_idx + 1:
            use_two_row_header = True
            logger.info(
                "header_two_row | name_row=%s date_row=%s base_date_str=%s",
                name_row_idx,
                date_header_row_idx,
                base_date_str,
            )
        elif name_row_idx is not None and date_header_row_idx == name_row_idx:
            logger.info("header_row_found | row_idx=%s base_date_str=%s", header_row_idx, base_date_str)
        else:
            logger.info("header_row_found | row_idx=%s (date_row_only) base_date_str=%s", header_row_idx, base_date_str)

        if use_two_row_header:
            df = pd.read_excel(file_path, header=[header_row_idx - 2, header_row_idx - 1])
        else:
            df = pd.read_excel(file_path, header=header_row_idx - 1)

        # 展平 MultiIndex 列名：取第二层（日期 2,3,…,六,日），无则取第一层（姓名、考勤组等），统一为 str
        raw_columns = df.columns
        if hasattr(raw_columns, "levels"):
            column_names = []
            for c in raw_columns:
                if isinstance(c, tuple):
                    second = c[-1] if len(c) > 1 else c[0]
                    first = c[0]
                    if second is None or (isinstance(second, float) and pd.isna(second)):
                        column_names.append(str(first) if first is not None else "")
                    elif str(second).strip() in ("", "nan") or "Unnamed" in str(second):
                        column_names.append(str(first) if first is not None else "")
                    else:
                        column_names.append(str(second))
                else:
                    column_names.append(str(c) if c is not None else "")
            df.columns = column_names
        else:
            column_names = [str(c) if c is not None else "" for c in df.columns]
        logger.info(
            "dataframe_loaded | header_row=%s columns_count=%s column_names=%s",
            header_row_idx,
            len(column_names),
            column_names[:30] if len(column_names) > 30 else column_names,
        )

        info_keywords = ["姓名", "考勤组", "部门", "工号", "职位", "考勤规则", "排班", "班次"]
        weekday_keywords = ["一", "二", "三", "四", "五", "六", "日", "天"]
        date_columns = []

        for col in column_names:
            col_str = str(col).strip()
            if any(keyword in col_str for keyword in info_keywords):
                continue
            if "Unnamed" in col_str or not col_str or col_str == "nan" or col_str == "":
                continue
            if "报表生成" in col_str or "生成时间" in col_str:
                continue
            if any(day in col_str for day in weekday_keywords):
                date_columns.append(col)
            elif col_str.isdigit() and len(col_str) <= 2:
                date_columns.append(col)

        if not date_columns:
            logger.error(
                "excel_core_fail | reason=no_date_columns column_names=%s "
                "（未识别到日期列：需列名包含一/二/…/日/天或1-2位数字）",
                column_names[:40] if len(column_names) > 40 else column_names,
            )
            return None, []

        first_date_col_index = column_names.index(date_columns[0]) if date_columns else 0
        logger.info(
            "date_columns_resolved | count=%s first_index=%s date_columns=%s",
            len(date_columns),
            first_date_col_index,
            date_columns,
        )

        night_records: List[Dict] = []
        rows_to_delete: List[int] = []
        total_rows = len(df)

        for idx, row in df.iterrows():
            excel_row = idx + header_row_idx + 1
            try:
                # 姓名包含“离职”的员工：整行删除并跳过后续处理
                if "姓名" in column_names:
                    try:
                        name_val = row.get("姓名", "")
                        name_str = "" if pd.isna(name_val) else str(name_val)
                        if "离职" in name_str:
                            rows_to_delete.append(excel_row)
                            continue
                    except Exception:
                        pass

                for date_col in date_columns:
                    header_value = date_col
                    col_idx = column_names.index(date_col)
                    date_str = parse_date_from_header(header_value, base_date_str, col_idx, first_date_col_index)
                    if not date_str:
                        date_str_raw = str(header_value).strip()
                        date_match = re.search(r"(\d{4}-\d{2}-\d{2})", date_str_raw)
                        if date_match:
                            date_str = date_match.group(1)
                        else:
                            continue

                    time_value = row[date_col]
                    if pd.isna(time_value) or time_value == "":
                        continue

                    time_str = str(time_value).strip()
                    if "休息" in time_str:
                        if is_holiday(date_str, custom_holidays):
                            col_idx_excel = column_names.index(date_col) + 1
                            cell = ws.cell(row=excel_row, column=col_idx_excel)
                            if cell.font:
                                cell.font = Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    underline=cell.font.underline,
                                    color="FFFF0000",
                                )
                            else:
                                cell.font = RED_FONT
                        continue

                    # 先做单元格内容清洗，再按原有逻辑标红与统计
                    cleaned_value = clean_daily_punch_cell(time_value, cutoff_time=clean_cutoff_time)
                    if cleaned_value is None or cleaned_value == "":
                        continue
                    if cleaned_value != time_str:
                        col_idx_excel = column_names.index(date_col) + 1
                        ws.cell(row=excel_row, column=col_idx_excel).value = cleaned_value
                    time_value_for_logic = cleaned_value

                    last_time = extract_last_time(time_value_for_logic)
                    if last_time is None:
                        continue

                    should_highlight = False

                    if is_holiday(date_str, custom_holidays):
                        should_highlight = True

                    if last_time >= late_time or last_time < early_time:
                        should_highlight = True

                    try:
                        night_flag = last_time < early_time
                        if night_flag:
                            name = row.get("姓名", "") if "姓名" in column_names else ""
                            emp_id = row.get("工号", "") if "工号" in column_names else ""
                            dept = row.get("部门", "") if "部门" in column_names else ""
                            group = row.get("考勤组", "") if "考勤组" in column_names else ""

                            weekday_label = ""
                            try:
                                _dt_obj = _dt.strptime(date_str, "%Y-%m-%d")
                                weekday_map = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
                                weekday_label = weekday_map[_dt_obj.weekday()]
                            except Exception:
                                pass

                            night_records.append(
                                {
                                    "name": str(name),
                                    "emp_id": str(emp_id),
                                    "dept": str(dept),
                                    "group": str(group),
                                    "date": date_str,
                                    "weekday": weekday_label,
                                    "last_time": last_time.strftime("%H:%M"),
                                    "all_times": str(time_value_for_logic),
                                    "is_holiday": bool(is_holiday(date_str, custom_holidays)),
                                }
                            )
                    except Exception:
                        pass

                    if should_highlight:
                        col_idx_excel = column_names.index(date_col) + 1
                        cell = ws.cell(row=excel_row, column=col_idx_excel)
                        if cell.font:
                            cell.font = Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                underline=cell.font.underline,
                                color="FFFF0000",
                            )
                        else:
                            cell.font = RED_FONT

            except Exception as cell_err:
                logger.warning(
                    "row_process_exception | excel_row=%s df_idx=%s error=%s",
                    excel_row,
                    idx,
                    cell_err,
                    exc_info=True,
                )
                continue

        # 统一删除离职员工行（倒序删除避免行号偏移）
        if rows_to_delete:
            logger.info("deleting_resigned_rows | count=%s rows=%s", len(rows_to_delete), sorted(set(rows_to_delete)))
            for r in sorted(set(rows_to_delete), reverse=True):
                try:
                    ws.delete_rows(r, 1)
                except Exception as del_err:
                    logger.warning("delete_row_failed | row=%s error=%s", r, del_err)

        dir_name, base_name = os.path.split(file_path)
        name_root, _ext = os.path.splitext(base_name)
        if name_root.endswith("_标红"):
            new_base = f"{name_root}.xlsx"
        else:
            new_base = f"{name_root}_标红.xlsx"
        output_path = os.path.join(dir_name, new_base)
        logger.info("saving_workbook | output_path=%s night_records=%s total_rows_processed=%s", output_path, len(night_records), total_rows)
        wb.save(output_path)

        logger.info("excel_core_success | output_path=%s night_records=%s", output_path, len(night_records))
        return output_path, night_records

    except Exception as e:
        logger.exception(
            "excel_core_exception | path=%s error=%s traceback=%s",
            file_path,
            e,
            traceback.format_exc(),
        )
        return None, []


